import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\wngud\Downloads\한국투자증권_오픈API_전체문서_20260217_030000.xlsx',
    read_only=True
)

out = open('kis_api_overseas.txt', 'w', encoding='utf-8')

# Extract overseas stock API details  
overseas_sheets = [s for s in wb.sheetnames if '해외주식' in s and '선물' not in s]
out.write(f"=== 해외주식 관련 API ({len(overseas_sheets)}개) ===\n\n")

for name in overseas_sheets:
    ws = wb[name]
    out.write(f"\n{'='*80}\n")
    out.write(f"SHEET: {name}\n")
    out.write(f"{'='*80}\n")
    for row in ws.iter_rows(max_row=60, values_only=True):
        cleaned = [str(c)[:120] if c else '' for c in row]
        out.write(f"  {cleaned}\n")

# Also get order-related APIs
out.write(f"\n\n{'#'*80}\n")
out.write(f"=== 주식주문(현금) 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['주식주문(현금)']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

# Get 주식현재가 시세 API  
out.write(f"\n\n{'#'*80}\n")
out.write(f"=== 주식현재가 시세 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['주식현재가 시세']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

# Get 주식현재가 일자별
out.write(f"\n\n{'#'*80}\n")  
out.write(f"=== 주식현재가 일자별 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['주식현재가 일자별']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

# Get balance checking API
out.write(f"\n\n{'#'*80}\n")
out.write(f"=== 주식잔고조회 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['주식잔고조회']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

# Get 거래량순위 API  
out.write(f"\n\n{'#'*80}\n")
out.write(f"=== 거래량순위 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['거래량순위']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

# Get 주식일별분봉조회
out.write(f"\n\n{'#'*80}\n")
out.write(f"=== 주식일별분봉조회 상세 ===\n")
out.write(f"{'#'*80}\n")
ws = wb['주식일별분봉조회']
for row in ws.iter_rows(max_row=80, values_only=True):
    cleaned = [str(c)[:120] if c else '' for c in row]
    out.write(f"  {cleaned}\n")

out.close()
wb.close()
print("Done! Check kis_api_overseas.txt")
