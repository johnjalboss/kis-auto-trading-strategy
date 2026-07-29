import openpyxl
import sys

wb = openpyxl.load_workbook(
    r'C:\Users\wngud\Downloads\한국투자증권_오픈API_전체문서_20260217_030000.xlsx',
    read_only=True
)

out = open('kis_api_summary.txt', 'w', encoding='utf-8')

out.write(f"Total sheets: {len(wb.sheetnames)}\n")
out.write(f"Sheet names: {wb.sheetnames}\n\n")

for name in wb.sheetnames:
    ws = wb[name]
    out.write(f"\n{'='*80}\n")
    out.write(f"SHEET: {name}\n")
    out.write(f"{'='*80}\n")
    count = 0
    for row in ws.iter_rows(max_row=30, values_only=True):
        cleaned = [str(c)[:100] if c else '' for c in row]
        out.write(f"  {cleaned}\n")
        count += 1
    out.write(f"  ... ({count} rows shown)\n")

out.close()
wb.close()
print("Done! Saved to kis_api_summary.txt")
